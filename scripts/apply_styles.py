"""Apply FBM Excel formatting standard to a workbook via openpyxl.

Cross-platform helper for the fbm-excel skill. Use this in cloud environments
(Claude.ai) or anywhere Excel COM isn't available.

Usage:
    from apply_styles import inject_fbm_styles, FBM
    from openpyxl import load_workbook

    wb = load_workbook('target.xlsx')
    inject_fbm_styles(wb)

    ws = wb.active
    ws['B2'].style = 'FBM Title'
    ws['B5'].style = 'FBM Header'
    ws['C6'].style = 'FBM Input $'
    wb.save('target.xlsx')
"""
from openpyxl.styles import (
    NamedStyle, Font, PatternFill, Alignment, Border, Side, Color
)
from openpyxl.formatting.rule import FormulaRule


class FBM:
    """FBM brand constants — colors, fonts, number formats."""

    # Brand colors (ARGB hex strings for openpyxl, no '#' prefix)
    NAVY       = '093254'
    BURGUNDY   = '6A1831'
    SLATE_BLUE = '4E7087'
    CHARCOAL   = '636466'
    FOREST     = '005E34'
    SAGE       = '708573'
    LIGHT_GREY = 'E7E6E6'
    BAND_GREY  = 'F7F7F7'
    WHITE      = 'FFFFFF'
    BLACK      = '000000'

    # Cell text colors (convention)
    INPUT_BLUE      = '0000FF'  # hardcoded inputs
    FORMULA_BLK     = '000000'  # formulas
    LINK_GREEN      = '008000'  # cross-sheet links
    EXTERNAL_RED    = 'FF0000'  # external file links
    ASSUMPTION_FILL = 'FFFF00'  # yellow fill for key assumptions

    # Tab colors
    TAB_BLUE  = '1F4E79'  # Inputs
    TAB_GREEN = '005E34'  # Outputs
    TAB_GRAY  = 'A5A5A5'  # Reference
    TAB_NAVY  = '093254'  # Cover

    # Conditional formatting palette
    GOOD_FONT,    GOOD_FILL    = '006100', 'C6EFCE'
    NEUTRAL_FONT, NEUTRAL_FILL = '9C5700', 'FFEB9C'
    BAD_FONT,     BAD_FILL     = '9C0006', 'FFC7CE'

    # Number format strings (single standard each)
    FMT_CURRENCY     = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    FMT_CURRENCY2    = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    FMT_NUMBER       = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
    FMT_PERCENT      = '0.0%;(0.0%);"-"'
    FMT_MULTIPLE     = '0.0"x"'
    FMT_DATE         = 'mm/dd/yyyy'
    FMT_YEAR         = '@'
    FMT_RATIO        = '0.00'
    FMT_BPS          = '0" bps"'
    FMT_SHARE_COUNT  = '#,##0'
    FMT_SHARE_PRICE  = '$#,##0.00'

    # Variance / delta formats — "positive good, negative bad"
    FMT_VARIANCE     = '+#,##0;(#,##0);"-"'
    FMT_VARIANCE_DOL = '+$#,##0;($#,##0);"-"'
    FMT_VARIANCE_PCT = '+0.0%;(0.0%);"-"'
    FMT_VARIANCE_BPS = '+0" bps";(0)" bps";"-"'

    # Fonts
    HEADER_FONT_NAME = 'Rockwell'   # fallback to Calibri Bold if unavailable
    BODY_FONT_NAME   = 'Calibri'
    BODY_FONT_SIZE   = 9
    HEADER_FONT_SIZE = 11
    TITLE_FONT_SIZE  = 14
    KPI_BIG_SIZE     = 24

    # Standard row heights (points)
    ROW_HEIGHT_TITLE        = 24
    ROW_HEIGHT_SUBTITLE     = 16
    ROW_HEIGHT_UNITS        = 16
    ROW_HEIGHT_HEADER       = 30
    ROW_HEIGHT_DATA         = 16
    ROW_HEIGHT_TOTAL_BREATH = 8

    # Print
    PRINT_SCALE_FLOOR = 60  # percent; reorganize before shrinking further


def _color(hex_no_hash: str) -> Color:
    """openpyxl Color from a 6-char ARGB hex string."""
    return Color(rgb='FF' + hex_no_hash)


def _make_style(name: str, **kwargs) -> NamedStyle:
    """Build a NamedStyle with the given attributes."""
    s = NamedStyle(name=name)
    if 'font' in kwargs:
        s.font = kwargs['font']
    if 'fill' in kwargs:
        s.fill = kwargs['fill']
    if 'alignment' in kwargs:
        s.alignment = kwargs['alignment']
    if 'border' in kwargs:
        s.border = kwargs['border']
    if 'number_format' in kwargs:
        s.number_format = kwargs['number_format']
    return s


def inject_fbm_styles(wb) -> None:
    """Add all FBM named styles to a workbook.

    Safe to call multiple times — skips styles that already exist.
    """
    existing = set(wb.named_styles)

    def add(style: NamedStyle) -> None:
        if style.name not in existing:
            wb.add_named_style(style)

    body = lambda **kw: Font(name=FBM.BODY_FONT_NAME, size=FBM.BODY_FONT_SIZE, **kw)

    # Title (Rockwell 14 navy)
    add(_make_style('FBM Title',
        font=Font(name=FBM.HEADER_FONT_NAME, size=FBM.TITLE_FONT_SIZE,
                  bold=True, color=_color(FBM.NAVY))))

    # Subtitle (italic charcoal, no wrap)
    add(_make_style('FBM Subtitle',
        font=body(italic=True, color=_color(FBM.CHARCOAL)),
        alignment=Alignment(horizontal='left', vertical='center', wrap_text=False)))

    # Units row (italic charcoal, centered, no wrap)
    add(_make_style('FBM Units',
        font=body(italic=True, color=_color(FBM.CHARCOAL)),
        alignment=Alignment(horizontal='center', vertical='center', wrap_text=False)))

    # Header (navy fill, white bold, centered)
    add(_make_style('FBM Header',
        font=Font(name=FBM.BODY_FONT_NAME, size=FBM.HEADER_FONT_SIZE,
                  bold=True, color=_color(FBM.WHITE)),
        fill=PatternFill('solid', fgColor=FBM.NAVY),
        alignment=Alignment(horizontal='center', vertical='center'),
        border=Border(*(Side(style='thin') for _ in range(4)))))

    # Subheader (light grey fill, navy bold)
    add(_make_style('FBM Subheader',
        font=Font(name=FBM.BODY_FONT_NAME, size=10, bold=True,
                  color=_color(FBM.NAVY)),
        fill=PatternFill('solid', fgColor=FBM.LIGHT_GREY),
        alignment=Alignment(horizontal='left'),
        border=Border(*(Side(style='thin') for _ in range(4)))))

    # Input styles (blue text)
    add(_make_style('FBM Input',
        font=body(color=_color(FBM.INPUT_BLUE)),
        number_format=FBM.FMT_NUMBER))
    add(_make_style('FBM Input $',
        font=body(color=_color(FBM.INPUT_BLUE)),
        number_format=FBM.FMT_CURRENCY))
    add(_make_style('FBM Input %',
        font=body(color=_color(FBM.INPUT_BLUE)),
        number_format=FBM.FMT_PERCENT))

    # Formula styles (black text)
    add(_make_style('FBM Formula',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_NUMBER))
    add(_make_style('FBM Formula $',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_CURRENCY))
    add(_make_style('FBM Formula %',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_PERCENT))

    # Variance styles — "positive good, negative bad"
    add(_make_style('FBM Variance',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_VARIANCE))
    add(_make_style('FBM Variance $',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_VARIANCE_DOL))
    add(_make_style('FBM Variance %',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_VARIANCE_PCT))
    add(_make_style('FBM Variance bps',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_VARIANCE_BPS))

    # Cross-sheet link (green text)
    add(_make_style('FBM Link',
        font=body(color=_color(FBM.LINK_GREEN)),
        number_format=FBM.FMT_NUMBER))

    # External link (red text)
    add(_make_style('FBM External',
        font=body(color=_color(FBM.EXTERNAL_RED)),
        number_format=FBM.FMT_NUMBER))

    # Key assumption (yellow fill, blue bold)
    add(_make_style('FBM Assumption',
        font=body(color=_color(FBM.INPUT_BLUE), bold=True),
        fill=PatternFill('solid', fgColor=FBM.ASSUMPTION_FILL),
        number_format='General'))

    # Total row (top single + bottom double border)
    add(_make_style('FBM Total',
        font=body(color=_color(FBM.FORMULA_BLK), bold=True),
        border=Border(top=Side(style='thin'), bottom=Side(style='double')),
        number_format=FBM.FMT_NUMBER))

    # Date
    add(_make_style('FBM Date',
        font=body(color=_color(FBM.FORMULA_BLK)),
        number_format=FBM.FMT_DATE))

    # Year (text, centered)
    add(_make_style('FBM Year',
        font=body(color=_color(FBM.FORMULA_BLK)),
        alignment=Alignment(horizontal='center'),
        number_format=FBM.FMT_YEAR))

    # Banding (alternating data-row fill)
    add(_make_style('FBM Band',
        font=body(color=_color(FBM.FORMULA_BLK)),
        fill=PatternFill('solid', fgColor=FBM.BAND_GREY)))

    # KPI Big (24pt Rockwell bold navy, centered)
    add(_make_style('FBM KPI Big',
        font=Font(name=FBM.HEADER_FONT_NAME, size=FBM.KPI_BIG_SIZE,
                  bold=True, color=_color(FBM.NAVY)),
        alignment=Alignment(horizontal='center', vertical='center')))

    # KPI Label (9pt italic charcoal, centered)
    add(_make_style('FBM KPI Label',
        font=body(italic=True, color=_color(FBM.CHARCOAL)),
        alignment=Alignment(horizontal='center', vertical='center')))


def apply_standard_layout(
    ws,
    header_row: int = 5,
    title_text: str = '',
    subtitle_text: str = '',
    units_text: str = '',
    units_range: str = '',
    label_col: str = 'B',
    with_banding: bool = True,
    band_rows: int = 40,
) -> None:
    """Apply standard FBM layout to a worksheet:
      - Col A = 5 (gutter)
      - Col B = 30 (labels)
      - Other cols = 15
      - Title in B2 with FBM Title style
      - Optional subtitle in B3 (italic charcoal, wrap OFF, overflow right)
      - Optional units row in row 4 (italic charcoal, centered) across units_range
      - Standard row heights applied (24/16/16/30/16)
      - Freeze panes computed from header_row and label_col
      - Optional banding on alternating data rows (rows header_row+1 .. +band_rows)
    """
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for col_letter in 'CDEFGHIJ':
        ws.column_dimensions[col_letter].width = 15

    if title_text:
        ws['B2'] = title_text
        ws['B2'].style = 'FBM Title'

    if subtitle_text:
        set_subtitle(ws, subtitle_text)

    if units_text:
        set_units(ws, units_text, cell_range=units_range or 'C4:J4')

    apply_row_heights(
        ws,
        header_row=header_row,
        has_subtitle=bool(subtitle_text),
        has_units=bool(units_text),
    )

    ws.freeze_panes = _freeze_anchor(header_row, label_col)

    if with_banding:
        first_data_row = header_row + 1
        last_data_row = first_data_row + band_rows - 1
        apply_banding(ws, f'B{first_data_row}:J{last_data_row}')


def set_subtitle(ws, text: str, row: int = 3, col: str = 'B') -> None:
    """Write an italic charcoal subtitle that does NOT wrap.

    Wrapping a long sentence inside the 30-width label column inflates the
    row height. This helper forces wrap_text=False and left alignment so
    long text overflows into the empty cells to the right.
    """
    cell = ws[f'{col}{row}']
    cell.value = text
    cell.style = 'FBM Subtitle'


def set_units(ws, text: str, cell_range: str = 'C4:J4') -> None:
    """Write a centered italic units label (e.g. '$ thousands', '%', 'bps')
    across the data columns of row 4. Merges the range and applies FBM Units.
    """
    ws.merge_cells(cell_range)
    top_left = cell_range.split(':')[0]
    cell = ws[top_left]
    cell.value = text
    cell.style = 'FBM Units'


def apply_banding(ws, cell_range: str, band_color: str | None = None) -> None:
    """Apply alternating-row banding via conditional formatting.

    Default band color is FBM.BAND_GREY (#F7F7F7). Uses MOD(ROW(),2)=0 so even
    rows get the band. Apply once per data block (not the whole sheet).
    """
    fill = PatternFill('solid', fgColor=band_color or FBM.BAND_GREY)
    rule = FormulaRule(formula=['MOD(ROW(),2)=0'], fill=fill)
    ws.conditional_formatting.add(cell_range, rule)


def set_indent(cell, level: int = 1) -> None:
    """Indent a label cell via Alignment(indent=level), never leading spaces.

    Sort-safe, copy-safe, machine-readable. Preserves existing horizontal/
    vertical alignment and wrap_text.
    """
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=existing.horizontal or 'left',
        vertical=existing.vertical or 'center',
        wrap_text=existing.wrap_text,
        indent=level,
    )


def apply_row_heights(
    ws,
    header_row: int = 5,
    has_subtitle: bool = True,
    has_units: bool = False,
) -> None:
    """Apply the FBM standard row heights.

    Title row (2), subtitle (3), units (4), header (header_row), data
    (header_row+1 .. +20) all get the locked heights from FBM.ROW_HEIGHT_*.
    """
    ws.row_dimensions[2].height = FBM.ROW_HEIGHT_TITLE
    if has_subtitle:
        ws.row_dimensions[3].height = FBM.ROW_HEIGHT_SUBTITLE
    if has_units:
        ws.row_dimensions[4].height = FBM.ROW_HEIGHT_UNITS
    ws.row_dimensions[header_row].height = FBM.ROW_HEIGHT_HEADER
    for r in range(header_row + 1, header_row + 21):
        ws.row_dimensions[r].height = FBM.ROW_HEIGHT_DATA


def apply_kpi_card(ws, big_cell: str, label_cell: str, value, label: str) -> None:
    """Write a KPI card: big number on top, small label below.

    big_cell:   e.g. 'C6' — gets FBM KPI Big
    label_cell: e.g. 'C7' — gets FBM KPI Label
    """
    ws[big_cell] = value
    ws[big_cell].style = 'FBM KPI Big'
    ws[label_cell] = label
    ws[label_cell].style = 'FBM KPI Label'


def hide_gridlines(ws) -> None:
    """Hide gridlines (Cover and Output convention)."""
    ws.sheet_view.showGridLines = False


def disable_pivot_autofit(wb) -> int:
    """For every pivot table in the workbook: disable autofit-on-refresh and
    enable preserve-formatting-on-refresh.

    Maps to the PivotTable Options checkboxes:
      - "Autofit column widths on update" -> OFF (useAutoFormatting=False)
      - "Preserve cell formatting on update" -> ON (preserveFormatting=True)

    Returns the count of pivot tables touched.
    """
    count = 0
    for ws in wb.worksheets:
        for pivot in getattr(ws, '_pivots', []):
            pivot.useAutoFormatting = False
            pivot.preserveFormatting = True
            count += 1
    return count


def set_workbook_properties(
    wb,
    title: str = '',
    subject: str = '',
    author: str = 'Andrew Kim',
    company: str = 'Foundation Building Materials',
    category: str = 'Operations FP&A',
) -> None:
    """Populate File -> Info -> Properties. Surfaces in DMS and SharePoint.

    Note: openpyxl's DocumentProperties exposes core OOXML fields (title,
    subject, creator, category, keywords) but not Company. Company is in
    docProps/app.xml which openpyxl writes separately; setting it reliably
    requires Excel COM (see build-template.ps1) or a custom property. We
    add it as a custom property as a best-effort fallback so it appears in
    File -> Info -> Show All Properties -> Custom.
    """
    props = wb.properties
    if title:
        props.title = title
    if subject:
        props.subject = subject
    if author:
        props.creator = author
        props.lastModifiedBy = author
    if category:
        props.category = category
    if company:
        try:
            from openpyxl.packaging.custom import StringProperty
            wb.custom_doc_props.append(StringProperty(name='Company', value=company))
        except Exception:
            pass


def _freeze_anchor(header_row: int, label_col: str) -> str:
    """Return the freeze-pane anchor: one row below header, one col right of label_col."""
    n = 0
    for ch in label_col.upper():
        n = n * 26 + (ord(ch) - ord('A') + 1)
    n += 1
    next_col = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        next_col = chr(r + ord('A')) + next_col
    return f'{next_col}{header_row + 1}'


def set_tab_color(ws, role: str) -> None:
    """Set worksheet tab color per FBM convention.

    role: 'input', 'output', 'reference', or 'cover'
    """
    colors = {
        'input':     FBM.TAB_BLUE,
        'output':    FBM.TAB_GREEN,
        'reference': FBM.TAB_GRAY,
        'cover':     FBM.TAB_NAVY,
    }
    if role not in colors:
        raise ValueError(f"role must be one of {list(colors)}")
    ws.sheet_properties.tabColor = colors[role]


if __name__ == '__main__':
    # Self-test: build a demo workbook exercising every helper
    from openpyxl import Workbook
    wb = Workbook()
    inject_fbm_styles(wb)
    set_workbook_properties(wb, title='FBM Style Demo', subject='Self-test')

    ws = wb.active
    ws.title = 'Demo'
    apply_standard_layout(
        ws,
        title_text='FBM Style Demo',
        subtitle_text='Demo of subtitle, units row, banding, variance formats, and KPI card.',
        units_text='$ thousands',
        units_range='C4:F4',
    )
    set_tab_color(ws, 'output')
    hide_gridlines(ws)

    # Header row
    for col, label in enumerate(['Item', 'FY2026', 'FY2027', 'FY2028', 'Fav/(Unfav) vs Plan'], start=2):
        cell = ws.cell(row=5, column=col, value=label)
        cell.style = 'FBM Header'

    # Sample input row
    ws['B6'] = 'Revenue'
    for col, val in enumerate([10_000, 10_500, 11_025], start=3):
        cell = ws.cell(row=6, column=col, value=val)
        cell.style = 'FBM Input $'
    ws['F6'] = 250
    ws['F6'].style = 'FBM Variance $'

    # Sub-item with indent
    ws['B7'] = 'Online channel'
    set_indent(ws['B7'], level=1)

    # Total row with blank breathing-room row above
    ws.row_dimensions[8].height = FBM.ROW_HEIGHT_TOTAL_BREATH
    ws['B9'] = 'Total revenue'
    for col_letter in ['C', 'D', 'E']:
        ws[f'{col_letter}9'] = f'={col_letter}6'
        ws[f'{col_letter}9'].style = 'FBM Total'
        ws[f'{col_letter}9'].number_format = FBM.FMT_CURRENCY

    # KPI card
    apply_kpi_card(ws, 'H6', 'H7', 23.5, 'Gross Margin %')

    wb.save('fbm_demo.xlsx')
    print('Saved fbm_demo.xlsx')
